#!/usr/bin/env python3
import os
import sys
import argparse
import datetime
import subprocess
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import re

# 设置中文字体首选和 CSS 样式
CSS_VARIABLES = """
:root {
  /* 顶级字体栈变量：全量强制首选本地苹方SC字体，确保长图整体字形统一与高清度 */
  --font-sans: "PingFang SC", -apple-system, BlinkMacSystemFont, sans-serif;
  --font-mono: "PingFang SC", -apple-system, BlinkMacSystemFont, monospace;

  --primary: #0f172a;       /* slate-900 */
  --primary-light: #1e293b; /* slate-800 */
  --accent: #3b82f6;        /* blue-500 鲜活高亮蓝 */
  --accent-light: rgba(59, 130, 246, 0.08);
  --bg-main: #ebf3fe;       /* 清澈亮丽的天蓝白色 */
  --bg-card: rgba(255, 255, 255, 0.82); /* 带有半透明毛玻璃感 */
  --text-main: #334155;     /* slate-700 */
  --text-dark: #0f172a;     /* slate-900 */
  --text-muted: #64748b;    /* slate-500 */
  --border-light: rgba(59, 130, 246, 0.08); /* 微微带蓝的精致细线 */
  
  /* 统一圆角黄金比例 */
  --radius-lg: 24px;
  --radius-md: 16px;
  --radius-sm: 8px;
  
  /* 活力爆棚的状态标签配色 */
  --badge-yes-bg: linear-gradient(135deg, rgba(59, 130, 246, 0.08), rgba(37, 99, 235, 0.03));
  --badge-yes-text: #2563eb;
  --badge-no-bg: linear-gradient(135deg, rgba(100, 116, 139, 0.05), rgba(71, 85, 105, 0.02));
  --badge-no-text: #64748b;
  
  /* 兵种标签统一为高对比高活力拼色 */
  --badge-af-bg: linear-gradient(135deg, rgba(16, 185, 129, 0.09), rgba(5, 150, 105, 0.03));   /* 安服：薄荷绿 */
  --badge-af-text: #059669;
  --badge-xf-bg: linear-gradient(135deg, rgba(249, 115, 22, 0.09), rgba(234, 88, 12, 0.03));   /* 行销：明橙 */
  --badge-xf-text: #ea580c;
  --badge-jf-bg: linear-gradient(135deg, rgba(99, 102, 241, 0.09), rgba(79, 70, 229, 0.03));   /* 技服：极光靛紫 */
  --badge-jf-text: #4f46e5;
  --badge-other-bg: linear-gradient(135deg, rgba(100, 116, 139, 0.08), rgba(71, 85, 105, 0.03));
  --badge-other-text: #475569;
  
  /* 晶莹环境立体投影 */
  --shadow-card: 0 10px 30px -5px rgba(59, 130, 246, 0.04), 0 4px 12px -2px rgba(0, 0, 0, 0.01);
  --shadow-stats: 0 16px 36px -12px rgba(59, 130, 246, 0.09);
}

/* 隐藏滚动条，避免极长长图产生右侧垂直滚动条占用排版宽度并遮挡卡片边框 */
::-webkit-scrollbar {
  display: none !important;
  width: 0 !important;
  height: 0 !important;
}

/* 全局字体渲染平滑度优化与高级等宽数字防抖，使长图上的数字在纵向完美对齐，字形温润清爽 */
* {
  -webkit-font-smoothing: antialiased;
  -moz-osx-font-smoothing: grayscale;
  text-rendering: optimizeLegibility;
  font-variant-numeric: tabular-nums;
  font-feature-settings: "kern" 1;
  text-shadow: 0 0 1px rgba(0,0,0,0.01); /* 消除部分无头浏览器渲染时的像素发虚 */
}
"""

LOGO_SVG = """<svg id="_图层_2" data-name="图层 2" xmlns="http://www.w3.org/2000/svg" viewBox="0 0 143.11 55.51">
  <defs>
    <style>
      .cls-1 {
        fill: #c70019;
      }
      .cls-2 {
        fill: #08287f;
      }
    </style>
  </defs>
  <g id="_图层_1-2" data-name="图层 1">
    <g>
      <g>
        <path class="cls-1" d="M90.65,49.13c-2.96,0-5.37-2.41-5.37-5.37s2.41-5.37,5.37-5.37c1.89,0,3.66,1.01,4.62,2.65,.08,.13,.1,.3,.06,.45-.04,.15-.14,.28-.28,.36l-8.21,4.75-.59-1.02,7.66-4.44c-.79-.98-2-1.57-3.27-1.57-2.31,0-4.19,1.88-4.19,4.19s1.88,4.19,4.19,4.19c1.5,0,2.9-.82,3.65-2.13l1.03,.58c-.95,1.68-2.75,2.73-4.67,2.73Z"/>
        <g>
          <rect class="cls-1" x="123.99" y="38.88" width="1.18" height="10.25"/>
          <path class="cls-1" d="M125.46,36.22c0,.49-.4,.88-.88,.88s-.88-.4-.88-.88,.39-.88,.88-.88,.88,.4,.88,.88Z"/>
        </g>
        <path class="cls-1" d="M101.74,49.13c-2.96,0-5.37-2.41-5.37-5.37s2.41-5.37,5.37-5.37c2.29,0,4.34,1.46,5.08,3.64l-1.12,.38c-.58-1.7-2.17-2.84-3.97-2.84-2.31,0-4.19,1.88-4.19,4.19s1.88,4.19,4.19,4.19c1.74,0,3.28-1.05,3.91-2.68l1.1,.42c-.81,2.08-2.77,3.43-5.01,3.43Z"/>
        <g>
          <path class="cls-1" d="M112.26,49.13c-2.6,0-4.72-2.12-4.72-4.72h1.18c0,1.95,1.59,3.54,3.54,3.54s3.54-1.59,3.54-3.54h1.18c0,2.6-2.12,4.72-4.71,4.72Z"/>
          <rect class="cls-1" x="115.8" y="38.65" width="1.18" height="10.25"/>
          <rect class="cls-1" x="107.54" y="38.65" width="1.18" height="5.76"/>
        </g>
        <g>
          <path class="cls-1" d="M119.51,43.41h-1.18c0-2.5,2.03-4.53,4.53-4.53v1.18c-1.85,0-3.35,1.5-3.35,3.35Z"/>
          <rect class="cls-1" x="118.33" y="38.88" width="1.18" height="10.25"/>
        </g>
        <g>
          <path class="cls-1" d="M131.98,49.13c-2.02,0-3.66-1.64-3.66-3.66h1.18c0,1.37,1.11,2.48,2.48,2.48v1.18Z"/>
          <rect class="cls-1" x="128.32" y="35.29" width="1.18" height="10.54"/>
          <rect class="cls-1" x="126.29" y="38.68" width="5.69" height="1.18"/>
        </g>
        <path class="cls-1" d="M80.69,49.13c-1.73,0-3.23-.74-4.11-2.04l.97-.66c.66,.97,1.81,1.53,3.14,1.53,.86,0,2.86-.22,2.86-2.25,0-1.03-1.59-1.39-2.95-1.39-.91,0-3.7,0-3.7-2.75,0-2.25,1.94-3.25,3.86-3.25,1.57,0,3.07,.78,3.73,1.95l-1.02,.58c-.45-.79-1.56-1.35-2.7-1.35-.45,0-2.68,.1-2.68,2.08,0,.97,.42,1.57,2.52,1.57,2.55,0,4.13,.98,4.13,2.57,0,2.08-1.59,3.43-4.04,3.43Z"/>
        <polygon class="cls-1" points="135.16 53.92 136.43 53.92 138.16 49.83 138.69 48.61 142.96 38.58 141.67 38.58 138.05 47.1 134.42 38.58 133.14 38.58 137.41 48.61 135.16 53.92"/>
        <g>
          <path class="cls-1" d="M52.47,48.85h-1.09l4.22-9.77h.97l4.2,9.77h-1.1l-1.08-2.57h-5.05l-1.06,2.57Zm1.44-3.48h4.31l-2.15-5.19-2.15,5.19Z"/>
          <path class="cls-1" d="M61.19,49.13h-1.71l-1.08-2.57h-4.68l-1.06,2.57h-1.7l4.46-10.33h1.33l4.44,10.33Zm-1.34-.55h.5l-3.96-9.22h-.6l-3.99,9.22h.49l1.06-2.57h5.42l1.08,2.57Zm-1.22-2.93h-5.14l2.57-6.19,2.57,6.19Zm-4.31-.55h3.48l-1.74-4.19-1.74,4.19Z"/>
        </g>
        <g>
          <path class="cls-1" d="M42.35,39.07h3.25c2.44,0,5.14,1.63,5.14,4.89s-2.69,4.89-5.14,4.89h-3.25v-9.77Zm.99,8.89h1.97c2.94,0,4.36-1.99,4.36-4s-1.42-4-4.36-4h-1.97v8.01Z"/>
          <path class="cls-1" d="M45.59,49.13h-3.52v-10.33h3.52c2.62,0,5.41,1.81,5.41,5.17s-2.79,5.16-5.41,5.16Zm-2.97-.55h2.97c2.41,0,4.86-1.58,4.86-4.61s-2.44-4.61-4.86-4.61h-2.97v9.22Zm2.69-.33h-2.25v-8.56h2.25c3.19,0,4.64,2.22,4.64,4.28s-1.45,4.28-4.64,4.28Zm-1.7-.55h1.7c2.82,0,4.09-1.87,4.09-3.73s-1.26-3.73-4.09-3.73h-1.7v7.45Z"/>
        </g>
        <g>
          <path class="cls-1" d="M67.87,40.4c-.41-.59-1.09-.98-2.02-.98s-2.08,.58-2.08,1.84,1.04,1.56,2.13,1.92c1.38,.46,2.9,.9,2.9,2.86s-1.64,2.82-3.19,2.82c-1.19,0-2.37-.48-3.08-1.45l.86-.63c.43,.69,1.22,1.23,2.24,1.23s2.17-.63,2.17-1.89c0-1.37-1.15-1.71-2.39-2.1-1.3-.4-2.62-.97-2.62-2.73,0-1.89,1.68-2.69,3.12-2.69,1.31,0,2.29,.55,2.76,1.19l-.79,.63Z"/>
          <path class="cls-1" d="M65.61,49.13c-1.35,0-2.59-.59-3.3-1.56l-.16-.22,1.32-.98,.16,.26c.43,.69,1.17,1.1,2,1.1,.92,0,1.89-.57,1.89-1.61s-.82-1.41-2.16-1.82c-1.02-.31-2.85-.88-2.85-3.01s1.76-2.97,3.4-2.97c1.49,0,2.52,.67,2.98,1.3l.16,.21-1.22,.99-.17-.24c-.39-.56-1.01-.86-1.79-.86-.73,0-1.81,.41-1.81,1.56,0,.97,.71,1.25,1.88,1.64l.18,.06c1.32,.44,2.97,.98,2.97,3.08s-1.8,3.09-3.46,3.09Zm-2.68-1.68c.63,.71,1.61,1.12,2.68,1.12,1.45,0,2.91-.79,2.91-2.54s-1.26-2.12-2.59-2.56l-.18-.06c-1.06-.35-2.26-.74-2.26-2.16s1.19-2.11,2.36-2.11c.84,0,1.56,.31,2.06,.87l.35-.28c-.44-.45-1.25-.87-2.37-.87-1.37,0-2.84,.76-2.84,2.42s1.2,2.09,2.43,2.47c1.29,.4,2.58,.8,2.58,2.36,0,1.41-1.26,2.17-2.44,2.17-.92,0-1.76-.41-2.3-1.11l-.4,.29Z"/>
        </g>
        <rect class="cls-1" x="70.7" y="43.23" width="4.38" height="1.39"/>
      </g>
      <g>
        <path class="cls-2" d="M59.98,29.06c-.37-.08-.73-.19-1.06-.33l-1.66-.77,4.18-1.94c.84-.37,1.55-.97,2.05-1.73,.5-.77,.77-1.66,.77-2.58v-3.34h1.24v-2.87h-1.24v-1.48h-2.87v1.48h-15.08v-1.48h-2.87v1.48h-1.29v2.87h1.29v3.34c0,.92,.27,1.81,.77,2.57,.5,.77,1.2,1.36,2.02,1.72l4.21,1.96-1.62,.75c-.35,.15-.72,.27-1.09,.35-.38,.08-.76,.12-1.14,.12h-4.52v2.87h4.52c.58,0,1.16-.06,1.73-.18,.58-.12,1.13-.29,1.69-.54l3.85-1.79,3.88,1.8c.53,.23,1.09,.41,1.66,.52,.57,.12,1.15,.18,1.73,.18h4.41v-2.87h-4.41c-.38,0-.77-.04-1.15-.12Zm-13.66-10.69h15.08v3.34c0,.36-.1,.7-.3,1-.2,.3-.48,.53-.83,.69l-6.41,2.98-6.44-2.99c-.33-.14-.61-.38-.8-.67-.2-.3-.3-.65-.3-1v-3.34Z"/>
        <polygon class="cls-2" points="44.99 12.34 62.58 12.34 62.58 13.48 65.45 13.48 65.45 9.47 55.27 9.47 55.27 8.61 52.4 8.61 52.4 9.47 42.13 9.47 42.13 13.48 44.99 13.48 44.99 12.34"/>
        <polygon class="cls-2" points="99.53 8.66 96.52 8.66 94.56 13.32 93.56 13.32 93.56 15.9 95.06 15.9 95.06 32.03 97.92 32.03 97.92 12.49 99.53 8.66"/>
        <polygon class="cls-2" points="117.56 9.52 109.93 9.52 109.93 8.66 107.06 8.66 107.06 9.52 99.53 9.52 99.53 12.39 117.56 12.39 117.56 9.52"/>
        <rect class="cls-2" x="99.53" y="13.68" width="18.03" height="2.87"/>
        <rect class="cls-2" x="99.53" y="17.83" width="18.03" height="2.87"/>
        <path class="cls-2" d="M99.46,32.04h13.95c1.11,0,2.16-.43,2.95-1.22,.79-.79,1.22-1.84,1.22-2.95v-6.2h-18.12v10.37Zm15.26-4.17c0,.35-.14,.68-.38,.92-.25,.25-.58,.38-.93,.38h-11.08v-4.64h12.39v3.33Z"/>
        <rect class="cls-2" x="74.77" y="8.82" width="16.71" height="2.66"/>
        <rect class="cls-2" x="74.77" y="29.39" width="16.71" height="2.66"/>
        <path class="cls-2" d="M91.36,23.84V12.5h-16.55v15.41h12.48c1.09,0,2.11-.42,2.88-1.19,.77-.77,1.19-1.79,1.19-2.88Zm-13.9-2.39h11.24v2.39c0,.38-.15,.73-.41,1-.27,.27-.62,.41-1,.41h-9.83v-3.8Zm11.24-2.66h-11.24v-3.64h11.24v3.64Z"/>
        <polygon class="cls-2" points="73.77 8.61 71.12 8.61 71.12 11.94 70.55 11.94 70.55 8.61 67.89 8.61 67.89 30.69 70.55 30.69 70.55 14.6 71.12 14.6 71.12 31.99 73.77 31.99 73.77 14.6 74.4 14.6 74.4 11.94 73.77 11.94 73.77 8.61"/>
        <path class="cls-2" d="M127.29,29.17c-.14-.14-.22-.33-.22-.52v-4.01h-2.66v4.01c0,.91,.35,1.76,1,2.4,.64,.64,1.49,1,2.4,1h10.87v-2.66h-10.87c-.2,0-.38-.08-.52-.22Z"/>
        <path class="cls-2" d="M141.55,22.5c.77-.77,1.19-1.79,1.19-2.88V9.69h-9.79v-1.08h-2.66v1.08h-9.79v14h18.17c1.09,0,2.11-.42,2.88-1.19Zm-1.46-9.04h-16.93v-1.11h16.93v1.11Zm0,2.66v1.03h-16.93v-1.03h16.93Zm-16.93,3.69h16.92c-.04,.31-.17,.59-.4,.81-.27,.27-.62,.41-1,.41h-15.52v-1.23Z"/>
        <path class="cls-2" d="M120.94,29.11s-.01,.13-.08,.2c-.08,.08-.17,.08-.2,.08h-.86v2.66h.86c.79,0,1.52-.31,2.08-.86,.55-.55,.86-1.29,.86-2.08v-4.47h-2.65v4.47Z"/>
        <path class="cls-2" d="M142.36,29.39s-.13-.01-.2-.08c-.07-.07-.08-.16-.08-.2v-4.47h-2.66v4.47c0,.78,.31,1.52,.86,2.08,.55,.56,1.29,.86,2.08,.86h.75v-2.66h-.75Z"/>
        <path class="cls-2" d="M128.01,24.64v.97c0,.52,.13,1.04,.42,1.47,.49,.74,1.28,1.15,2.14,1.15h8.09v-2.57h-7.78c-.17,0-.3-.13-.3-.29v-.74h-2.57Z"/>
      </g>
      <g>
        <path class="cls-1" d="M37.07,.04c-3.67,1.33-7.28,2.79-10.83,4.37l.06-1.22c0-.06,0-.12,0-.18-.02-.23-.21-.41-.45-.41-.03,0-.05,0-.08,0-.03,0-.06,.01-.09,.03-.03,0-.06,.02-.08,.03-1.06,.35-2.21,.66-3.28,.96-.4,.11-.8,.22-1.2,.32-.43,.11-.85,.22-1.28,.33-.39,.1-.77,.19-1.16,.28-.45,.11-.89,.21-1.34,.31-.38,.08-.76,.17-1.14,.25-.46,.1-.91,.19-1.37,.28-.38,.08-.77,.15-1.15,.22-.47,.09-.94,.17-1.41,.25-.37,.06-.74,.12-1.1,.18-.49,.08-.98,.15-1.47,.22-.36,.05-.72,.1-1.09,.15-.49,.07-.98,.13-1.47,.19-.38,.05-.76,.09-1.15,.13-.47,.05-.94,.1-1.4,.14-.42,.04-.83,.08-1.25,.11-.44,.04-.88,.07-1.33,.1-.31,.02-.62,.04-.94,.06-.2,.01-.41,.03-.61,.04-.04,0-.18,0-.29,.01-.04,0-.07,0-.11,.02-.05,.02-.08,.07-.08,.13,0,.08,.05,.13,.13,.15,.15,.04,.37,.07,.45,.08,5.12,.97,10.7,1.55,15.91,1.55h0c-.18,.1-.36,.19-.55,.29-1.52,.84-2.98,1.76-4.37,2.76-.05,.04-.11,.08-.16,.11-1.29,.94-2.42,2.08-3.35,3.38-1.39,1.94-2.32,4.03-2.32,6.43,0,.87,.12,1.75,.36,2.59,.02,.08,.05,.16,.07,.23,.07,.21,.14,.42,.22,.62,.06,.14,.2,.24,.37,.24,.21,0,.38-.16,.4-.36,0-.05,0-.11,0-.17,0-.16,0-.32,.01-.49,0-.12,.02-.24,.03-.36,.05-.5,.14-1,.27-1.47,.39-1.42,1.22-2.89,2.13-4.03,.69-.86,1.46-1.6,2.31-2.33,6.12-5.29,15.53-9.07,23.13-11.34,.33-.11,.57-.23,.74-.57l.05-.12,.27-.59,1.53-3.37s.02-.03,.02-.05c.02-.04,.02-.09,.02-.14,0-.24-.2-.44-.44-.44-.04,0-.08,0-.12,.02-.02,0-.05,.01-.07,.02"/>
        <path class="cls-1" d="M33.61,7.24c-1,.31-2,.63-2.99,.98-3.07,1.07-6.06,2.25-8.98,3.68-5.73,2.82-12.56,7.33-13.04,12.93-.11,1.35,.14,2.72,.77,3.93,.27,.51,.59,.99,.97,1.42,.1,.11,.19,.22,.3,.32,.26,.26,.55,.51,.85,.73h0c.05,.03,.12,.05,.19,.05,.19,0,.35-.16,.35-.35,0-.05-.01-.11-.03-.15h0c-.22-.45-.4-.94-.52-1.44-.22-.9-.25-1.84-.11-2.76,.13-.81,.4-1.61,.8-2.32,3.14-6.03,11.97-9.3,17.7-11.31,.72-.25,1.45-.5,2.19-.73,.01,0,.03-.01,.04-.01,.14-.05,.26-.12,.36-.22,.1-.1,.14-.18,.2-.31l1.59-3.5,.03-.07c.05-.11,.08-.19,.08-.31,0-.32-.26-.58-.58-.58-.05,0-.11,0-.15,.02,0,0,0,0-.01,0"/>
        <path class="cls-1" d="M8.19,42.53c-.61,.21-1.22,.42-1.84,.62-.14,.05-.29,.09-.43,.14-.18,.07-.4,.25-.48,.43l-1.63,3.59-.06,.13h0c-.03,.07-.04,.14-.04,.21,0,.33,.27,.6,.6,.6,.05,0,.1,0,.15-.02,1.22-.38,2.42-.77,3.62-1.2,.66-.23,1.32-.48,1.97-.73,2.07-.79,4.28-1.7,6.28-2.67,2.66-1.29,5.2-2.87,7.54-4.69,.62-.47,1.23-.97,1.81-1.49,1.56-1.37,2.75-3.15,3.39-5.18,.38-1.13,.49-2.31,.3-3.49-.29-1.76-1.25-3.35-2.65-4.43-.11-.09-.22-.19-.37-.19-.19,0-.34,.15-.34,.34,0,.04,0,.07,.02,.1,.01,.04,.02,.06,.04,.1,.45,.93,.7,1.97,.7,3.06,0,1.21-.3,2.38-.89,3.43-.93,1.79-2.26,3.34-3.85,4.54-2.05,1.55-4.36,2.83-6.71,3.92-2.31,1.07-4.73,2.01-7.13,2.85"/>
        <path class="cls-1" d="M31.75,30.17s-.04-.11-.06-.16c-.07-.2-.18-.36-.41-.36-.14,0-.26,.07-.33,.18-.09,.13-.07,.29-.07,.44,0,.18-.01,.36-.03,.55-.1,1.27-.43,2.3-.97,3.44-.67,1.34-1.57,2.57-2.66,3.6-3.77,3.64-7.76,5.76-12.39,7.88-3.82,1.75-7.78,3.25-11.87,4.47-.41,.12-.6,.28-.79,.7-.08,.19-.17,.38-.26,.57l-1.45,3.2c-.06,.12-.14,.28-.14,.41,0,.23,.19,.42,.41,.42,.12,0,.35-.1,.47-.14,6.26-2.28,14.24-5.48,20.67-9.15,3.63-2.07,6.79-4.06,8.85-7.54,.85-1.44,1.43-2.89,1.56-4.57,.06-.75,.03-1.52-.09-2.26-.09-.57-.24-1.13-.43-1.67"/>
      </g>
    </g>
  </g>
</svg>"""

def parse_args():
    parser = argparse.ArgumentParser(description="将工作计划 Excel 转换为精美的手机查看图片")
    parser.add_argument("-i", "--input", default="/Users/wangjunhua/Desktop/00.工作资料/agi/工作区/工作安排/(技术管理)技术申请.xlsx", help="输入的 Excel 文件路径（默认为 /Users/wangjunhua/Desktop/00.工作资料/agi/工作区/工作安排/(技术管理)技术申请.xlsx）")
    parser.add_argument("-o", "--output-dir", help="输出文件夹路径（默认为输入文件同级目录）")
    return parser.parse_args()

def find_chrome_binary():
    """
    智能寻找系统中的 Chrome 或 Chromium 可执行文件路径，提高跨平台运行稳定性
    """
    # 1. 尝试从环境变量获取
    env_path = os.environ.get("CHROME_PATH")
    if env_path and os.path.exists(env_path):
        return env_path
        
    # 2. 针对 macOS 的默认路径
    mac_paths = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
        "/Applications/Chromium.app/Contents/MacOS/Chromium"
    ]
    for p in mac_paths:
        if os.path.exists(p):
            return p
            
    # 3. 针对 Linux/Unix 系统的 path 查找
    import shutil
    for cmd in ["google-chrome", "google-chrome-stable", "chromium", "chromium-browser", "microsoft-edge"]:
        path = shutil.which(cmd)
        if path:
            return path
            
    return None

def simplify_plan_time(time_str):
    """
    智能合并和简化“计划时间”段文本，支持去重、排序及区间连续合并（如：周一下午至周五下午）
    """
    if not isinstance(time_str, str) or not time_str.strip() or time_str == 'nan':
        return ""
        
    # 用常见分隔符切分
    parts = re.split(r'[,，、;\s\n\+]+', time_str)
    slots_found = set()
    has_unrecognized = False
    
    days = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    day_map = {d: i for i, d in enumerate(days)}
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        # 匹配 "周X全天"
        match_all = re.match(r'^(周[一二三四五六日])全天$', part)
        if match_all:
            d = match_all.group(1)
            d_idx = day_map[d]
            slots_found.add(d_idx * 2)     # 上午
            slots_found.add(d_idx * 2 + 1) # 下午
            continue
            
        # 匹配 "周X上午" 或 "周X下午"
        match_half = re.match(r'^(周[一二三四五六日])(上午|下午)$', part)
        if match_half:
            d = match_half.group(1)
            t = match_half.group(2)
            d_idx = day_map[d]
            t_idx = 0 if t == '上午' else 1
            slots_found.add(d_idx * 2 + t_idx)
            continue
            
        has_unrecognized = True
        
    # 如果识别出 slot，但存在无法识别的内容（如特定备注等），为防破坏信息，保持原样
    if not slots_found or has_unrecognized:
        return time_str
        
    sorted_slots = sorted(list(slots_found))
    
    # 合并连续的时间槽
    ranges = []
    if sorted_slots:
        start = sorted_slots[0]
        prev = sorted_slots[0]
        for val in sorted_slots[1:]:
            if val == prev + 1:
                prev = val
            else:
                ranges.append((start, prev))
                start = val
                prev = val
        ranges.append((start, prev))
        
    # 辅助还原槽为文字
    def slot_to_str(idx):
        d_idx = idx // 2
        t_idx = idx % 2
        d_str = days[d_idx]
        t_str = '上午' if t_idx == 0 else '下午'
        return d_str, t_str
        
    range_strs = []
    for r_start, r_end in ranges:
        parts_list = []
        s = r_start
        e = r_end
        
        # 1. 剥离头部半天（若起始于下午，则该天单独构成“X下午”并剥离）
        if s % 2 == 1:
            d_str, t_str = slot_to_str(s)
            parts_list.append(f"{d_str}{t_str}")
            s += 1
            
        # 2. 剥离尾部半天（若结束于上午，则该天单独构成“X上午”并剥离）
        tail_part = None
        if e % 2 == 0:
            d_str, t_str = slot_to_str(e)
            tail_part = f"{d_str}{t_str}"
            e -= 1
            
        # 3. 合并中间的整天部分
        if s <= e:
            ds, _ = slot_to_str(s)
            de, _ = slot_to_str(e)
            if ds == de:
                parts_list.append(f"{ds}全天")
            else:
                parts_list.append(f"{ds}至{de}全天")
                
        # 补回尾部半天
        if tail_part:
            parts_list.append(tail_part)
            
        # 如果这个区间只有头部半天和尾部半天（中间没有整天），说明是跨日的连续半天，用“至”连接；否则用“、”连接
        if len(parts_list) == 2 and (r_start % 2 == 1) and (r_end % 2 == 0):
            range_strs.append("至".join(parts_list))
        else:
            range_strs.append("、".join(parts_list))
            
    return "、".join(range_strs)

def clean_and_prepare_data(input_path):
    """
    读取并清洗 Excel 数据，按销售排序，新增序号，保留指定字段
    """
    if not os.path.exists(input_path):
        print(f"错误：输入文件不存在 {input_path}")
        sys.exit(1)
        
    df = pd.read_excel(input_path)
    
    # 筛选：该周周一日期是当前时间之后的
    if '该周周一日期' in df.columns:
        temp_dates = pd.to_datetime(df['该周周一日期'], errors='coerce')
        df = df[temp_dates > pd.Timestamp.now()].copy()
    
    # 1. 排序（按销售申请任务数降序、总工时数降序、姓名拼音升序排序，确保 Excel 和手机长图序号是完全一致且顺序递增的）
    if '销售' in df.columns:
        sales_counts = df['销售'].value_counts().to_dict()
        df['_task_count'] = df['销售'].map(lambda x: sales_counts.get(x, 0))
        
        # 计算每个销售的预计工时总和以作为第二排序键
        df['预计工时（h）'] = pd.to_numeric(df['预计工时（h）'], errors='coerce').fillna(0)
        sales_hours_dict = df.groupby('销售')['预计工时（h）'].sum().to_dict()
        df['_sales_hours'] = df['销售'].map(lambda x: sales_hours_dict.get(x, 0))
        
        try:
            from pypinyin import pinyin, Style
            def get_pinyin_key(name):
                p_list = pinyin(str(name), style=Style.NORMAL)
                return "".join([w[0] for w in p_list]).lower()
        except ImportError:
            def get_pinyin_key(name):
                return str(name).encode('gb18030', errors='ignore')
                
        df['_pinyin'] = df['销售'].map(get_pinyin_key)
        # 按任务数降序、总工时降序、姓名拼音升序排序
        df = df.sort_values(by=['_task_count', '_sales_hours', '_pinyin'], ascending=[False, False, True])
        df = df.drop(columns=['_task_count', '_sales_hours', '_pinyin'])
        df = df.reset_index(drop=True)
    else:
        df['销售'] = ''
        
    # 3. 填充 NaN 为空字符串，预计工时如果为 NaN 填 0 或空，保持数据干净
    df = df.fillna('')
    
    # 提报发起人别名映射归一化
    if '提报发起人' in df.columns:
        name_map = {
            '饭团Dai': '张进',
            '钱钱': '钱丽云',
            '华': '段振华'
        }
        df['提报发起人'] = df['提报发起人'].astype(str).str.strip().replace(name_map)
    
    # 4. 新增序号字段并放置在第一列，格式化为两位数
    df.insert(0, '序号', [f"{i:02d}" for i in range(1, len(df) + 1)])
    
    # 5. 保留指定字段并按顺序排列
    target_columns = ['序号', '所在月周', '销售', '提报发起人', '兵种', '客户名称', '工作类型', '是否支持', '支持人员', '计划时间', '预计工时（h）', '备注']
    
    # 容错：如果缺少某些字段，创建空列
    for col in target_columns:
        if col not in df.columns:
            df[col] = ''
            
    df = df[target_columns].copy()
    
    # 对计划时间进行智能合并简化
    if '计划时间' in df.columns:
        df['计划时间'] = df['计划时间'].astype(str).apply(simplify_plan_time)
        
    # 确保预计工时（h）列在空值或非数字时填充为 0
    df['预计工时（h）'] = pd.to_numeric(df['预计工时（h）'], errors='coerce').fillna(0)
    return df

def beautify_excel(file_path):
    """
    使用 openpyxl 对导出的 Excel 进行精美样式美化
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    font_family = "Microsoft YaHei"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
    
    data_font = Font(name=font_family, size=10)
    data_fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # slate-50
    data_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1") # slate-300
    border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # 格式化表头
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border
        
    # 格式化数据行
    for row_idx in range(2, ws.max_row + 1):
        fill = data_fill_even if row_idx % 2 == 0 else data_fill_odd
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.fill = fill
            cell.border = border
            
            # 对齐逻辑
            col_name = ws.cell(row=1, column=col_idx).value
            if col_name in ['序号', '所在月周', '销售', '提报发起人', '是否支持', '计划时间']:
                cell.alignment = align_center
            elif col_name in ['预计工时（h）']:
                cell.alignment = align_right
                # 尽量保证数字类型
                try:
                    if cell.value != '':
                        cell.value = float(cell.value)
                except ValueError:
                    pass
            else:
                cell.alignment = align_left
                
    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            cell_len = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in val)
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)
        
    # 行高
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        
    wb.save(file_path)

def crop_image_by_marker(image_path, marker_color=(15, 23, 42), tolerance=25):
    """
    从下往上扫描图片，寻找特定颜色的 marker 横线，进行高度裁剪以防多余留白 and 内容截断
    """
    import os
    if not os.path.exists(image_path):
        return False
    try:
        img = Image.open(image_path)
        width, height = img.size
        
        # 获取快速像素加载通道，避免在 Python 中多次调用 getpixel 方法的开销，提速数十倍
        pixels = img.load()
        
        # 从下往上每隔 2 行扫描一次，大幅提升效率
        # 限制最大扫描深度（向上最深扫描至整体的 20% 高度），防止在超长空白时发生误判并减少计算量
        max_scan_limit = max(200, height // 5)
        marker_y = None
        for y in range(height - 10, max_scan_limit, -2):
            match_count = 0
            # 在图像中间 30% 范围内采样，确保效率并避开边缘渐变
            sample_points = range(int(width * 0.35), int(width * 0.65), 6)
            for x in sample_points:
                pixel = pixels[x, y]
                # 兼容 RGB 和 RGBA 格式
                r, g, b = pixel[0], pixel[1], pixel[2]
                if (abs(r - marker_color[0]) <= tolerance and 
                    abs(g - marker_color[1]) <= tolerance and 
                    abs(b - marker_color[2]) <= tolerance):
                    match_count += 1
            
            # 如果这一行有 80% 的采样像素点符合颜色，说明找到了底部分割线
            if match_count > (len(sample_points) * 0.8):
                marker_y = y
                break
                
        if marker_y is not None:
            # 裁剪到 marker 线下方 10 像素处
            cropped_img = img.crop((0, 0, width, min(height, marker_y + 10)))
            cropped_img.save(image_path, optimize=True)
            print(f"智能图像处理成功：检测到结束标记 (y={marker_y})，长图高度自适应裁剪为 {marker_y + 10}px")
            return True
        else:
            print("提示：未能在图像中定位底部分割线，保持原图大小。")
            return False
    except Exception as e:
        print(f"图像裁剪处理出错：{e}")
        return False

def get_badge_class(val):
    if val == "是":
        return "badge-yes"
    elif val == "否":
        return "badge-no"
    elif "安服" in str(val):
        return "badge-af"
    elif "行销" in str(val):
        return "badge-xf"
    elif "技服" in str(val):
        return "badge-jf"
    else:
        return "badge-other"

def generate_table_html(df, stats):
    """
    生成精美大表格 HTML 字符串
    """
    rows_html = ""
    for _, row in df.iterrows():
        support_cls = get_badge_class(row['是否支持'])
        type_cls = get_badge_class(row['兵种'])
        
        rows_html += f"""
        <tr>
            <td class="text-center">{row['序号']}</td>
            <td class="text-center font-semibold">{row['所在月周']}</td>
            <td class="text-center font-semibold text-dark">{row['销售']}</td>
            <td class="text-center text-muted">{row['提报发起人']}</td>
            <td class="text-center"><span class="badge {type_cls}">{row['兵种']}</span></td>
            <td class="font-semibold text-dark">{row['客户名称']}</td>
            <td>{row['工作类型']}</td>
            <td class="text-center"><span class="badge {support_cls}">{row['是否支持']}</span></td>
            <td>{row['支持人员']}</td>
            <td class="text-center text-muted">{row['计划时间']}</td>
            <td class="text-right font-mono font-semibold">{row['预计工时（h）']}</td>
            <td class="text-muted">{row['备注']}</td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="utf-8">
        <!-- 使用本地系统预装的高清字体栈（PingFang SC/Microsoft YaHei/SF Pro），防网络超时导致字体模糊 -->
        <style>
            {CSS_VARIABLES}
            
            body {{
                font-family: var(--font-sans);
                background-color: var(--bg-main);
                color: var(--text-main);
                margin: 0;
                padding: 30px;
                width: 1200px;
                box-sizing: border-box;
            }}
            
            .container {{
                background-color: var(--bg-card);
                border-radius: 16px;
                box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.05), 0 4px 6px -4px rgba(0, 0, 0, 0.05);
                border: 1px solid var(--border-light);
                padding: 30px;
            }}
            
            /* 标题区域 */
            .header {{
                border-bottom: 2px solid var(--accent-light);
                padding-bottom: 20px;
                margin-bottom: 24px;
                display: flex;
                justify-content: space-between;
                align-items: flex-end;
            }}
            
            .header-title h1 {{
                font-size: 26px;
                color: var(--text-dark);
                margin: 0 0 8px 0;
                font-weight: 700;
                letter-spacing: -0.5px;
            }}
            
            .header-title p {{
                font-size: 14px;
                color: var(--text-muted);
                margin: 0;
            }}
            
            .header-logo {{
                font-size: 13px;
                background: linear-gradient(135deg, var(--accent), #1d4ed8);
                color: white;
                padding: 6px 12px;
                border-radius: 20px;
                font-weight: 600;
            }}
            
            /* 数据面板 */
            .dashboard {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 20px;
                margin-bottom: 30px;
            }}
            
            .card {{
                background: var(--bg-main);
                border: 1px solid var(--border-light);
                border-radius: 12px;
                padding: 16px 20px;
                box-sizing: border-box;
            }}
            
            .card-label {{
                font-size: 13px;
                color: var(--text-muted);
                margin-bottom: 6px;
                font-weight: 500;
            }}
            
            .card-value {{
                font-size: 24px;
                font-weight: 700;
                color: var(--text-dark);
            }}
            
            .card-value span {{
                font-size: 14px;
                font-weight: 500;
                color: var(--text-muted);
                margin-left: 4px;
            }}
            
            /* 表格设计 */
            table {{
                width: 100%;
                border-collapse: separate;
                border-spacing: 0;
                margin-top: 10px;
                border-radius: 8px;
                overflow: hidden;
                border: 1px solid var(--border-light);
            }}
            
            th {{
                background-color: var(--primary-light);
                color: white;
                font-weight: 600;
                font-size: 14px;
                padding: 14px 12px;
                text-align: left;
                border-bottom: 1px solid var(--border-color);
            }}
            
            td {{
                padding: 12px;
                font-size: 13.5px;
                border-bottom: 1px solid var(--border-light);
                color: var(--text-main);
                word-break: break-all;
                line-height: 1.5;
            }}
            
            tr:nth-child(even) td {{
                background-color: var(--bg-main);
            }}
            
            tr:last-child td {{
                border-bottom: none;
            }}
            
            /* 辅助对齐类 */
            .text-center {{ text-align: center; }}
            .text-right {{ text-align: right; }}
            .font-semibold {{ font-weight: 600; }}
            .font-mono {{ font-family: "PingFang SC", -apple-system, sans-serif; }}
            .text-dark {{ color: var(--text-dark); }}
            .text-muted {{ color: var(--text-muted); }}
            
            /* 徽章 Badge */
            .badge {{
                display: inline-block;
                padding: 3px 8px;
                border-radius: 6px;
                font-size: 12px;
                font-weight: 600;
            }}
            .badge-yes {{ background-color: var(--badge-yes-bg); color: var(--badge-yes-text); }}
            .badge-no {{ background-color: var(--badge-no-bg); color: var(--badge-no-text); }}
            .badge-af {{ background-color: var(--badge-af-bg); color: var(--badge-af-text); }}
            .badge-xf {{ background-color: var(--badge-xf-bg); color: var(--badge-xf-text); }}
            .badge-jf {{ background-color: var(--badge-jf-bg); color: var(--badge-jf-text); }}
            .badge-other {{ background-color: var(--badge-other-bg); color: var(--badge-other-text); }}
            
            /* 页脚 */
            .footer {{
                margin-top: 30px;
                text-align: center;
                font-size: 12px;
                color: var(--text-muted);
                border-top: 1px solid var(--border-light);
                padding-top: 15px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="dashboard">
                <div class="card">
                    <div class="card-label">申请总数</div>
                    <div class="card-value">{stats['total_count']}<span>项</span></div>
                </div>
                <div class="card">
                    <div class="card-label">预计总工时</div>
                    <div class="card-value">{stats['total_hours']}<span>小时</span></div>
                </div>
                <div class="card">
                    <div class="card-label">涉及客户数</div>
                    <div class="card-value">{stats['total_customers']}<span>家</span></div>
                </div>
                <div class="card">
                    <div class="card-label">未提交申请销售</div>
                    <div class="card-value">{stats['unapplied_sales_count']}<span>人</span></div>
                    <div class="card-sub" style="font-size: 11px; color: var(--text-muted); margin-top: 4px; line-height: 1.2; word-break: break-all;">名单：{stats['unapplied_sales_names'] or '无'}</div>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th class="text-center" style="width: 50px;">序号</th>
                        <th class="text-center" style="width: 90px;">所在月周</th>
                        <th class="text-center" style="width: 80px;">销售</th>
                        <th class="text-center" style="width: 90px;">提报发起人</th>
                        <th class="text-center" style="width: 70px;">兵种</th>
                        <th style="width: 160px;">客户名称</th>
                        <th style="width: 140px;">工作类型</th>
                        <th class="text-center" style="width: 80px;">是否支持</th>
                        <th style="width: 100px;">支持人员</th>
                        <th class="text-center" style="width: 110px;">计划时间</th>
                        <th class="text-right" style="width: 80px;">工时(h)</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
            
            <div class="footer">
                页面宽度已适配电脑端大图查看
            </div>
        </div>
    </body>
    </html>
    """
    return html

def get_sales_avatar_style(name, idx=0):
    """
    根据销售人员的排列顺序依次轮流分派莫兰迪色系渐变，钱丽云作为江西办唯一女销售固定粉色渐变，其余男销售按次序循环分派，彻底防同色相邻。
    """
    name_str = str(name).strip()
    if name_str == "钱丽云":
        # 钱丽云（女性销售）固定为温馨优雅的珊瑚粉渐变
        bg, fg = ("linear-gradient(135deg, #f472b6, #fb7185)", "#ffffff")
    else:
        # 其他男性销售在其余经典莫兰迪配色中按排列次序循环分配（排除粉色）
        palettes = [
            ("linear-gradient(135deg, #a78bfa, #c084fc)", "#ffffff"), # 薰衣草紫
            ("linear-gradient(135deg, #fb923c, #ffb703)", "#ffffff"), # 暖金橙
            ("linear-gradient(135deg, #60a5fa, #3b82f6)", "#ffffff"), # 经典海蓝
            ("linear-gradient(135deg, #34d399, #059669)", "#ffffff"), # 翡翠绿
            ("linear-gradient(135deg, #2ec4b6, #0f172a)", "#ffffff"), # 青碧绿
        ]
        color_idx = idx % len(palettes)
        bg, fg = palettes[color_idx]
        
    return f"background: {bg}; color: {fg}; box-shadow: 0 4px 10px rgba(0, 0, 0, 0.08);"

def generate_card_html(df, stats, week_info):
    """
    生成针对手机屏幕（竖屏）的精美卡片流 HTML 字符串，贯彻 taste-skill 殿堂级通透设计标准
    """
    # 增加空数据防御
    if df.empty:
        return f"""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                {CSS_VARIABLES}
                html {{
                    margin: 0;
                    padding: 0;
                    width: 480px;
                    max-width: 480px;
                    overflow-x: hidden;
                    box-sizing: border-box;
                }}
                body {{
                     font-family: "PingFang SC", sans-serif;
                    background-color: var(--bg-main);
                    color: var(--text-main);
                    margin: 0;
                    padding: 24px 12px 24px 12px;
                    width: 480px;
                    max-width: 480px;
                    box-sizing: border-box;
                    overflow-x: hidden;
                }}
                .container {{
                    background-color: var(--bg-card);
                    border-radius: var(--radius-lg);
                    box-shadow: var(--shadow-stats);
                    border: 1px solid var(--border-light);
                    padding: 60px 24px;
                    text-align: center;
                }}
                .empty-icon {{
                    width: 64px;
                    height: 64px;
                    color: var(--text-muted);
                    margin: 0 auto 16px auto;
                }}
                .empty-title {{
                    font-size: 18px;
                    font-weight: 700;
                    color: var(--text-dark);
                    margin-bottom: 8px;
                }}
                .empty-desc {{
                    font-size: 13px;
                    color: var(--text-muted);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <svg class="empty-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round">
                    <circle cx="12" cy="12" r="10"></circle>
                    <line x1="12" y1="8" x2="12" y2="12"></line>
                    <line x1="12" y1="16" x2="12.01" y2="16"></line>
                </svg>
                <div class="empty-title">暂无本周技术安排</div>
                <div class="empty-desc">导入的数据表格中未发现有效计划记录。</div>
                <!-- 智能裁剪定位标记线 -->
                <div style="height: 2px; background-color: #0f172a; margin-top: 40px;"></div>
            </div>
        </body>
        </html>
        """

    SALES_LIST = ["邓正春", "饶达琴", "段振华", "钱丽云", "詹文成", "伍斌", "吴刚", "张进", "贺欢"]
    all_sales = list(SALES_LIST)
    if '销售' in df.columns:
        excel_sales = [str(x).strip() for x in df['销售'].dropna().unique() if str(x).strip()]
        for s in excel_sales:
            if s not in all_sales:
                all_sales.append(s)
                
    # 构造每个销售的数据以及任务数、总工时以做排序
    sales_data_list = []
    for s_name in all_sales:
        if s_name in df['销售'].values:
            group = df[df['销售'] == s_name]
            task_count = len(group)
            try:
                sales_hours = pd.to_numeric(group['预计工时（h）'], errors='coerce').fillna(0).sum()
            except Exception:
                sales_hours = 0
        else:
            group = pd.DataFrame(columns=df.columns)
            task_count = 0
            sales_hours = 0
        sales_data_list.append((s_name, group, task_count, sales_hours))
        
    # 确保排序规则与 df 完全一致，让卡片上的序号顺次递增：
    # 1. 有任务的销售：根据他们在 df 中首次出现的索引进行排序（此时 df 已排好序：任务数降序、拼音升序）
    # 2. 0 任务的销售：排在后面，他们之间按拼音升序排序
    try:
        from pypinyin import pinyin, Style
        def get_pinyin_key(name):
            p_list = pinyin(str(name), style=Style.NORMAL)
            return "".join([w[0] for w in p_list]).lower()
    except ImportError:
        def get_pinyin_key(name):
            return str(name).encode('gb18030', errors='ignore')

    def get_sales_sort_key(item):
        s_name = item[0]
        if s_name in df['销售'].values:
            return (0, df[df['销售'] == s_name].index[0])
        else:
            return (1, get_pinyin_key(s_name))

    sales_data_list.sort(key=get_sales_sort_key)
    
    sections_html = ""
    male_color_counter = 0
    for sales_name, group, task_count, sales_hours in sales_data_list:
        sales_char = str(sales_name)[0] if sales_name else "销"
        name_str = str(sales_name).strip()
        if name_str == "钱丽云":
            avatar_style = get_sales_avatar_style(sales_name)
        else:
            avatar_style = get_sales_avatar_style(sales_name, idx=male_color_counter)
            male_color_counter += 1
            
        group_cards_html = ""
        if group.empty:
            group_cards_html = """
            <div class="empty-cards-notice">
                <svg class="empty-notice-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
                    <circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="12"/><line x1="12" y1="16" x2="12.01" y2="16"/>
                </svg>
                暂无本周技术申请
            </div>
            """
            
        for _, row in group.iterrows():
            type_cls = get_badge_class(row['兵种'])
            
            # 根据兵种生成极细矢量图标
            army_type = str(row['兵种']).strip()
            if "安服" in army_type:
                army_icon_html = f"""
                <span class="badge {type_cls}">
                    <svg class="badge-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/></svg>
                    {row['兵种']}
                </span>
                """
            elif "行销" in army_type:
                army_icon_html = f"""
                <span class="badge {type_cls}">
                    <svg class="badge-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><path d="m16.24 7.76-1.41 4.24-4.24 1.41 1.41-4.24 4.24-1.41Z"/></svg>
                    {row['兵种']}
                </span>
                """
            elif "技服" in army_type:
                army_icon_html = f"""
                <span class="badge {type_cls}">
                    <svg class="badge-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><path d="M14.7 6.3a1 1 0 0 0 0 1.4l1.6 1.6a1 1 0 0 0 1.4 0l3.77-3.77a6 6 0 0 1-7.94 7.94l-6.91 6.91a2.12 2.12 0 0 1-3-3l6.91-6.91a6 6 0 0 1 7.94-7.94l-3.76 3.76z"/></svg>
                    {row['兵种']}
                </span>
                """
            else:
                army_icon_html = f"""
                <span class="badge {type_cls}">
                    <svg class="badge-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polygon points="12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2"/></svg>
                    {row['兵种']}
                </span>
                """
            
            # 协作状态深度融合：合并是否支持与支持人员，展示精致的线稿头像徽章，并增加边框防褪色
            is_support = str(row['是否支持']).strip() == "是" and str(row['支持人员']).strip()
            if is_support:
                collab_html = f"""
                <span class="badge badge-yes">
                    <div class="avatar-stack">
                        <div class="stack-avatar"></div>
                        <div class="stack-avatar"></div>
                    </div>
                    {row['支持人员']}
                </span>
                """
            else:
                collab_html = '<span class="badge badge-no">否</span>'
            
            # 备注处理，如果为空则不显示备注区块，增加精致线稿便签小图标
            remark_html = ""
            if str(row['备注']).strip():
                remark_html = f"""
                <div class="remark-box">
                    <span class="field-label">
                        <svg class="badge-icon" style="vertical-align: -1px; margin-right: 4px;" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
                            <path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"></path>
                        </svg>备注
                    </span>
                    <span class="field-value text-muted remark-value">{row['备注']}</span>
                </div>
                """
            
            # 增加计划卡片的不对称聚焦律动（若为协作型项目，升级为 plan-card-accent 高亮聚焦卡）
            card_cls = "plan-card plan-card-accent" if is_support else "plan-card"
            
            group_cards_html += f"""
            <div class="{card_cls}">
                <!-- 殿堂级艺术感背景超大字号序号水印，绝对定位在右下角，立体穿插 -->
                <div class="card-watermark">{row['序号']}</div>
                
                <div class="card-header">
                    <!-- 左侧添加微渐变 Focus Anchor（视觉落脚蓝色聚焦条） -->
                    <div class="customer-name-wrapper">
                        <span class="customer-anchor"></span>
                        <span class="customer-name">{row['客户名称']}</span>
                    </div>
                    <div class="card-header-badges">
                        {army_icon_html}
                        <span class="badge badge-hours">
                            <svg class="badge-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                                <circle cx="12" cy="12" r="10"/>
                                <polyline points="12 6 12 12 16 14"/>
                            </svg>
                            {row['预计工时（h）']}小时
                        </span>
                    </div>
                </div>
                
                <div class="card-body">
                    <div class="card-grid">
                        <div class="grid-item-type">
                            <span class="field-label">
                                <svg class="badge-icon" style="vertical-align: -1.5px; margin-right: 3px;" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                                    <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
                                    <polyline points="14 2 14 8 20 8"/>
                                </svg>工作类型
                            </span>
                            <span class="field-value font-semibold text-dark">{row['工作类型']}</span>
                        </div>
                        <div class="grid-item-creator">
                            <span class="field-label">
                                <svg class="badge-icon" style="vertical-align: -1.5px; margin-right: 3px;" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                                    <path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/>
                                    <circle cx="12" cy="7" r="4"/>
                                </svg>发起人
                            </span>
                            <span class="field-value font-semibold text-dark">{row['提报发起人'] or '无'}</span>
                        </div>
                        <div class="grid-item-collab">
                            <span class="field-label">
                                <svg class="badge-icon" style="vertical-align: -1.5px; margin-right: 3px;" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                                    <path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/>
                                    <circle cx="9" cy="7" r="4"/>
                                    <path d="M23 21v-2a4 4 0 0 0-3-3.87"/>
                                    <path d="M16 3.13a4 4 0 0 1 0 7.75"/>
                                </svg>协作安排
                            </span>
                            <span class="field-value">{collab_html}</span>
                        </div>
                        <div class="grid-item-time">
                            <span class="field-label">
                                <svg class="badge-icon" style="vertical-align: -1.5px; margin-right: 3px;" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                                    <rect x="3" y="4" width="18" height="18" rx="2" ry="2"/>
                                    <line x1="16" y1="2" x2="16" y2="6"/>
                                    <line x1="8" y1="2" x2="8" y2="6"/>
                                    <line x1="3" y1="10" x2="21" y2="10"/>
                                </svg>计划时间
                            </span>
                            <span class="field-value text-dark">{row['计划时间']}</span>
                        </div>
                    </div>
                    
                    {remark_html}
                </div>
            </div>
            """
            
        if not group.empty and '提报发起人' in group.columns:
            m_count = len(group[group['提报发起人'] == name_str])
        else:
            m_count = 0
        tech_count = len(group) - m_count
        
        if len(group) == 0:
            task_info_str = "0项任务"
        else:
            if m_count == 0:
                detail_str = "全部为技术发起"
            elif tech_count == 0:
                detail_str = "全部为销售发起"
            else:
                detail_str = f"销售发起{m_count}项，技术发起{tech_count}项"
            task_info_str = f"{len(group)}项任务<span style=\"font-size: 10px; font-weight: 500; opacity: 0.85; margin-left: 2px;\">（{detail_str}）</span>"

        sections_html += f"""
        <div class="sales-section">
            <div class="sales-section-header">
                <div class="sales-info">
                    <span class="sales-avatar" style="{avatar_style}">{sales_char}</span>
                    <span class="sales-title">{sales_name}</span>
                </div>
                <div class="sales-meta">
                    <span class="meta-badge">{task_info_str}</span>
                    <span class="meta-badge hours-badge">{sales_hours:g}小时工时</span>
                </div>
            </div>
            <div class="sales-cards">
                {group_cards_html}
            </div>
        </div>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <!-- 使用本地系统预装的高清字体栈（PingFang SC/Microsoft YaHei/SF Pro），防网络超时导致字体模糊 -->
        <style>
            {CSS_VARIABLES}
            
            html {{
                margin: 0;
                padding: 0;
                width: 480px;
                max-width: 480px;
                overflow-x: hidden;
                box-sizing: border-box;
            }}
            
            body {{
                font-family: "PingFang SC", sans-serif;
                /* 引入重复弥散极光渐变，搭配超高级极淡蓝紫点阵背景，营造前沿数字画报视感 */
                background-image: 
                    radial-gradient(circle at 0% 0%, rgba(59, 130, 246, 0.18) 0%, transparent 60%),
                    radial-gradient(circle at 100% 25%, rgba(249, 115, 22, 0.15) 0%, transparent 55%),
                    radial-gradient(circle at 10% 70%, rgba(168, 85, 247, 0.15) 0%, transparent 50%),
                    radial-gradient(circle at 90% 90%, rgba(16, 185, 129, 0.13) 0%, transparent 45%),
                    radial-gradient(rgba(59, 130, 246, 0.035) 1.2px, transparent 0);
                background-size: 100% 100%, 100% 100%, 100% 100%, 100% 100%, 20px 20px;
                color: var(--text-main);
                margin: 0;
                padding: 18px 12px 30px 12px;
                width: 480px;
                max-width: 480px;
                box-sizing: border-box;
                overflow-x: hidden;
                -webkit-font-smoothing: antialiased;
                -moz-osx-font-smoothing: grayscale;
                text-rendering: optimizeLegibility;
            }}
            

            /* 手机专属头部（浅色通透毛玻璃画报风，极其明朗） */
            .header {{
                position: relative;
                background: rgba(255, 255, 255, 0.85);
                backdrop-filter: blur(20px);
                -webkit-backdrop-filter: blur(20px);
                color: var(--text-dark);
                border-radius: var(--radius-lg);
                padding: 16px 20px 12px 20px;
                margin-bottom: 12px;
                display: flex;
                flex-direction: column;
                align-items: center;
                gap: 2px;
                border: 1px solid rgba(255, 255, 255, 0.7);
                box-shadow: var(--shadow-stats);
                overflow: hidden;
            }}
            
            /* 对角扫光 Shimmer 效果 */
            .header::after {{
                content: "";
                position: absolute;
                top: 0;
                left: -150%;
                width: 100%;
                height: 100%;
                background: linear-gradient(
                    90deg,
                    transparent,
                    rgba(255, 255, 255, 0.25) 30%,
                    rgba(255, 255, 255, 0.45) 50%,
                    rgba(255, 255, 255, 0.25) 70%,
                    transparent
                );
                transform: skewX(-25deg);
                animation: shimmer 6s infinite ease-in-out;
            }}
            @keyframes shimmer {{
                0% {{ left: -150%; }}
                50% {{ left: -150%; }}
                100% {{ left: 150%; }}
            }}
            
            .header-sub {{
                font-size: 7px;
                color: var(--text-muted);
                letter-spacing: 0.18em;
                font-weight: 750;
                margin-top: -2px;
                opacity: 0.85;
            }}
            
            .header-logo-wrapper {{
                display: flex;
                justify-content: center;
                align-items: center;
                margin-bottom: 0px;
                width: 100%;
                z-index: 2;
                position: relative;
            }}
            
            .header-logo-wrapper svg {{
                height: 22px;
                width: auto;
            }}
            
            /* 精致的微蓝点状装饰线，倾注画报级质感 */
            .header::before {{
                content: "";
                position: absolute;
                inset: 0;
                background-image: radial-gradient(rgba(59, 130, 246, 0.08) 1.5px, transparent 0);
                background-size: 16px 16px;
                pointer-events: none;
                z-index: 1;
            }}
            
            .header-decor, .header h1 {{
                position: relative;
                z-index: 2;
            }}
            
            .header-decor {{
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 38px;
                height: 38px;
                background: linear-gradient(135deg, #3b82f6, #60a5fa);
                border-radius: 50%;
                color: white; 
                box-shadow: 0 4px 10px rgba(59, 130, 246, 0.2);
            }}
            
            .header-icon {{
                width: 18px;
                height: 18px;
            }}
            
            .header h1 {{
                font-size: 20px;
                margin: 0;
                font-weight: 800;
                letter-spacing: 0.04em; 
                color: var(--text-dark);
                text-align: center;
                background: linear-gradient(135deg, var(--text-dark), #1e293b);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }}
            
            /* 手机统计栏 Double-Bezel 双嵌套结构，炫彩宝石图标，注入满满的活力感 */
            .stats-bar-shell {{
                background-color: rgba(59, 130, 246, 0.04); 
                border-radius: var(--radius-lg);
                padding: 4px;
                margin-bottom: 24px;
                border: 1px solid rgba(59, 130, 246, 0.08);
            }}
            
            .stats-bar {{
                display: flex;
                justify-content: space-between;
                background: rgba(255, 255, 255, 0.85); 
                backdrop-filter: blur(12px);
                -webkit-backdrop-filter: blur(12px);
                border-radius: calc(var(--radius-lg) - 4px);
                padding: 14px 16px;
                box-shadow: var(--shadow-stats);
                border: 1px solid rgba(255, 255, 255, 0.6);
            }}
            
            .stat-item {{
                position: relative;
                flex: 1;
                min-width: 0;
                display: flex;
                align-items: flex-start;
                gap: 10px;
                padding: 6px 8px 8px 8px;
                background: rgba(255, 255, 255, 0.45);
                border: 1px solid rgba(255, 255, 255, 0.5);
                border-radius: var(--radius-sm);
                margin: 0 3px;
                box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.6);
            }}
            
            /* 底部 2.5px 高彩发光指示线条，引入霓虹阴影增强发光感 */
            .stat-item::after {{
                content: "";
                position: absolute;
                bottom: 0;
                left: 10px;
                right: 10px;
                height: 2.5px;
                border-radius: 1.5px;
            }}
            .stat-item-1::after {{ background-color: #3b82f6; box-shadow: 0 2px 8px rgba(59, 130, 246, 0.65); }}
            .stat-item-2::after {{ background-color: #f97316; box-shadow: 0 2px 8px rgba(249, 115, 22, 0.65); }}
            .stat-item-3::after {{ background-color: #a855f7; box-shadow: 0 2px 8px rgba(168, 85, 247, 0.65); }}
            
            /* 细分统计项宽度，第一二列收缩，为第三列预留宽敞空间以防标题和顿号姓名折行 */
            .stat-item-1 {{ flex: 0.82 !important; }}
            .stat-item-2 {{ flex: 0.85 !important; }}
            .stat-item-3 {{ flex: 1.33 !important; }}
            
            .stat-icon-wrapper {{
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 34px;
                height: 34px;
                border-radius: 10px;
                color: #ffffff !important;
                flex-shrink: 0;
                box-shadow: 0 4px 10px rgba(0, 0, 0, 0.04);
            }}
            
            .bg-blue-gradient {{ background: linear-gradient(135deg, #3b82f6, #60a5fa); }}
            .bg-orange-gradient {{ background: linear-gradient(135deg, #f97316, #facc15); }}
            .bg-purple-gradient {{ background: linear-gradient(135deg, #a855f7, #ec4899); }}
            
            .stat-icon {{
                width: 16px;
                height: 16px;
            }}
            
            .stat-info-block {{
                display: flex;
                flex-direction: column;
                gap: 2px;
            }}
            
            .stat-label {{
                font-size: 10px;
                color: var(--text-muted);
                text-transform: uppercase;
                font-weight: 700;
                letter-spacing: 0.03em;
                white-space: nowrap; /* 强制单行呈现，不折行 */
            }}
            
            /* 字重层级对比排版 */
            .stat-val {{
                font-size: 24px;
                font-weight: 900;
                color: var(--text-dark);
                font-family: var(--font-sans);
                line-height: 1;
                letter-spacing: -0.02em;
            }}
            
            .stat-item-1 .stat-val, .stat-item-2 .stat-val {{
                white-space: nowrap; /* 强制数值与单位在单行呈现，防折行 */
            }}
            
            .stat-val span {{
                font-size: 10px;
                font-weight: 500;
                color: var(--text-muted);
                margin-left: 2px;
                letter-spacing: 0;
            }}
            
            /* 分组模块 */
            .sales-section {{
                margin-bottom: 24px;
            }}
            
            .sales-section-header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                padding: 4px;
                margin-bottom: 12px;
                position: relative;
            }}
            
            /* 渐变虚线分栏引线，由中心向两端柔和淡出，极富光影质感 */
            .sales-section-header::after {{
                content: "";
                flex-grow: 1;
                height: 1px;
                background-image: linear-gradient(to right, transparent, rgba(59, 130, 246, 0.35) 20%, rgba(139, 92, 246, 0.35) 50%, rgba(59, 130, 246, 0.35) 80%, transparent);
                margin: 0 16px;
                opacity: 0.8;
            }}
            
            .sales-info {{
                display: flex;
                align-items: center;
            }}
            
            /* 销售人员 Avatar 升级为渐变潮流发光球 */
            .sales-avatar {{
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 28px;
                height: 28px;
                border-radius: 50%;
                margin-right: 10px;
                font-size: 13px;
                font-weight: 800;
                line-height: 1;
                transition: transform 0.2s ease;
            }}
            
            .sales-title {{
                font-size: 16px;
                font-weight: 700;
                color: var(--text-dark);
                letter-spacing: 0.02em;
            }}
            
            .sales-meta {{
                display: flex;
                gap: 6px;
                order: 3; 
            }}
            
            .meta-badge {{
                font-size: 11.5px;
                background-color: rgba(59, 130, 246, 0.08); 
                color: #2563eb;      
                padding: 4px 12px;
                border-radius: 20px;
                font-weight: 700;          
                border: 1px solid rgba(59, 130, 246, 0.15);
                box-shadow: 0 1px 2px rgba(59, 130, 246, 0.02);
            }}
            
            .hours-badge {{
                background-color: rgba(249, 115, 22, 0.08); 
                color: #ea580c;            
                border-color: rgba(249, 115, 22, 0.15);
            }}
            
            /* 卡片定义 - 高保真毛玻璃透光看板 */
            .sales-cards {{
                display: flex;
                flex-direction: column;
                gap: 8px;
            }}
            
            .plan-card {{
                position: relative;
                overflow: hidden;
                background: rgba(255, 255, 255, 0.8);
                backdrop-filter: blur(16px);
                -webkit-backdrop-filter: blur(16px);
                border: 1px solid rgba(255, 255, 255, 0.75); 
                border-radius: var(--radius-md); 
                padding: 16px 16px;
                box-shadow: 0 12px 28px -4px rgba(59, 130, 246, 0.04), 0 4px 12px -2px rgba(0, 0, 0, 0.015), inset 0 1.5px 0 rgba(255, 255, 255, 0.9);
                z-index: 2;
            }}
            
            .plan-card-accent {{
                background: rgba(255, 255, 255, 0.84);
                border: 1px solid rgba(59, 130, 246, 0.18); 
                box-shadow: 0 16px 36px -8px rgba(59, 130, 246, 0.08), 0 4px 16px -2px rgba(59, 130, 246, 0.03), inset 0 1.5px 0 rgba(255, 255, 255, 0.95);
            }}
            
            /* 霓虹发光边 */
            .plan-card::before {{
                content: "";
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                height: 2.5px;
                background: linear-gradient(to right, #3b82f6, #10b981);
                z-index: 3;
            }}
            
            .plan-card-accent::before {{
                content: "";
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                height: 3px;
                background: linear-gradient(to right, #8b5cf6, #ec4899);
                box-shadow: 0 1px 6px rgba(139, 92, 246, 0.4);
                z-index: 3;
            }}
            
            /* 隐约的精密网格点线肌理 */
            .plan-card::after {{
                content: "";
                position: absolute;
                inset: 0;
                background-image: 
                    linear-gradient(rgba(59, 130, 246, 0.012) 1px, transparent 1px),
                    linear-gradient(90deg, rgba(59, 130, 246, 0.012) 1px, transparent 1px);
                background-size: 24px 24px;
                pointer-events: none;
                z-index: 0;
            }}
            
            /* 卡片右下角背景序号水印 - 优雅的艺术空心轮廓字 */
            .card-watermark {{
                position: absolute;
                bottom: -6px;
                right: 14px;
                font-size: 80px; 
                font-weight: 900;
                font-style: italic; 
                color: transparent;
                -webkit-text-stroke: 1.5px rgba(59, 130, 246, 0.12);
                font-family: "SF Pro Display", "Helvetica Neue", Arial, sans-serif;
                line-height: 1;
                pointer-events: none;
                z-index: 1;
                transform: rotate(-3deg); 
                mix-blend-mode: multiply;
            }}
            
            /* 呼吸灯脉冲微光源 */
            .pulse-indicator {{
                position: absolute;
                top: 14px;
                right: 14px;
                display: flex;
                align-items: center;
                justify-content: center;
                z-index: 5;
            }}
            
            .pulse-dot {{
                width: 6px;
                height: 6px;
                border-radius: 50%;
                background-color: #8b5cf6;
                box-shadow: 0 0 8px #8b5cf6;
                position: relative;
            }}
            
            .pulse-dot::after {{
                content: "";
                position: absolute;
                width: 14px;
                height: 14px;
                border: 1px solid rgba(139, 92, 246, 0.5);
                border-radius: 50%;
                top: -5px;
                left: -5px;
                animation: pulse-ring 1.8s infinite cubic-bezier(0.25, 0, 0, 1);
            }}
            
            @keyframes pulse-ring {{
                0% {{
                    transform: scale(0.6);
                    opacity: 1;
                }}
                100% {{
                    transform: scale(1.6);
                    opacity: 0;
                }}
            }}
            
            .card-header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 8px;
                position: relative;
                z-index: 2;
            }}
            
            .customer-name-wrapper {{
                display: flex;
                align-items: center;
                gap: 8px;
                max-width: 70%;
            }}
            
            /* 视觉聚焦锚点（天蓝到薄荷绿的活力双色渐变，带霓虹微投影） */
            .customer-anchor {{
                display: inline-block;
                width: 4px;
                height: 16px;
                background: linear-gradient(to bottom, #3b82f6, #10b981);
                border-radius: 2.5px;
                flex-shrink: 0;
                box-shadow: 0 2px 8px rgba(59, 130, 246, 0.5);
            }}
            
            .plan-card-accent {{
                border-left-width: 1px; /* 移除侧边粗线条，统一使用顶部霓虹切角边 */
            }}
            
            .card-header-badges {{
                display: flex;
                gap: 6px;
                align-items: center;
            }}
            
            .customer-name {{
                font-size: 14px;
                font-weight: 700;
                color: var(--text-dark);
                line-height: 1.45;
                letter-spacing: 0.02em;
                word-break: break-all;
            }}
            
            .card-body {{
                display: flex;
                flex-direction: column;
                gap: 8px;
                position: relative;
                z-index: 2;
            }}
            
            .card-grid {{
                display: flex;
                justify-content: flex-start;
                gap: 16px;
            }}
            
            .grid-item {{
                flex: 1;
                display: flex;
                flex-direction: column;
                gap: 2px;
            }}
            
            .grid-item-type {{
                width: 85px;
                flex-shrink: 0;
                display: flex;
                flex-direction: column;
                gap: 2px;
            }}
            
            .grid-item-creator {{
                width: 80px;
                flex-shrink: 0;
                display: flex;
                flex-direction: column;
                gap: 2px;
            }}
            
            .grid-item-collab {{
                width: 95px;
                flex-shrink: 0;
                display: flex;
                flex-direction: column;
                gap: 2px;
            }}
            
            .grid-item-time {{
                flex-grow: 1;
                display: flex;
                flex-direction: column;
                gap: 2px;
            }}
            
            .field-label {{
                font-size: 11px;
                color: var(--text-muted);
                font-weight: 700;
                letter-spacing: 0.08em; 
                text-transform: uppercase;
                margin-bottom: 1px;
            }}
            
            .field-value {{
                font-size: 13px;
                color: var(--text-main);
                font-weight: 500; /* 加厚字重，确保截图不发虚 */
                line-height: 1.5;
                letter-spacing: 0.02em;
                word-break: break-all;
            }}
            
            .border-top {{
                border-top: 1px solid rgba(59, 130, 246, 0.06); 
                padding-top: 6px;
                margin-top: 2px;
            }}
            
            /* 升级为炫彩透光的备注栏，引入磨砂渐变和圆角悬浮竖条 */
            .remark-box {{
                position: relative;
                background: linear-gradient(135deg, rgba(59, 130, 246, 0.03), rgba(139, 92, 246, 0.01)); 
                padding: 10px 14px 10px 18px;
                border-radius: var(--radius-sm);
                margin-top: 10px;
                font-size: 12.5px;
                line-height: 1.6;
                word-break: break-all;
                border: 1px solid rgba(59, 130, 246, 0.04);
                z-index: 2;
            }}
            
            .remark-box::before {{
                content: "";
                position: absolute;
                top: 10px;
                bottom: 10px;
                left: 6px;
                width: 3px;
                background: linear-gradient(to bottom, #3b82f6, #8b5cf6);
                border-radius: 1.5px;
            }}
            
            .remark-value::first-letter {{
                font-size: 16px;
                font-weight: 800;
                color: var(--text-dark);
                float: left;
                margin-right: 2px;
                line-height: 1;
            }}
            
            /* 徽章，加宽字距，色彩更加清新饱满 */
            .badge {{
                display: inline-flex;
                align-items: center;
                padding: 4px 8px;
                border-radius: 6px;
                font-size: 11px;
                font-weight: 700;
                letter-spacing: 0.02em; 
                width: fit-content;
                line-height: 1;
                /* 晶莹亚克力质感 */
                backdrop-filter: blur(4px);
                -webkit-backdrop-filter: blur(4px);
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.02);
            }}
            .badge-yes {{ 
                background-color: var(--badge-yes-bg); 
                color: var(--badge-yes-text); 
                border: 1px solid rgba(59, 130, 246, 0.22);
                box-shadow: 0 1px 2px rgba(59, 130, 246, 0.03);
            }}
            .badge-no {{ 
                background-color: var(--badge-no-bg); 
                color: var(--badge-no-text); 
                border: 1px solid rgba(100, 116, 139, 0.15);
            }}
            .badge-af {{ 
                background-color: var(--badge-af-bg); 
                color: var(--badge-af-text); 
                border: 1px solid rgba(16, 185, 129, 0.25);
                box-shadow: 0 1px 2px rgba(16, 185, 129, 0.03);
            }}
            .badge-xf {{ 
                background-color: var(--badge-xf-bg); 
                color: var(--badge-xf-text); 
                border: 1px solid rgba(249, 115, 22, 0.25);
                box-shadow: 0 1px 2px rgba(249, 115, 22, 0.03);
            }}
            .badge-jf {{ 
                background-color: var(--badge-jf-bg); 
                color: var(--badge-jf-text); 
                border: 1px solid rgba(99, 102, 241, 0.25);
                box-shadow: 0 1px 2px rgba(99, 102, 241, 0.03);
            }}
            .badge-other {{ 
                background-color: var(--badge-other-bg); 
                color: var(--badge-other-text); 
                border: 1px solid rgba(100, 116, 139, 0.15);
            }}
            
            /* 工时胶囊：橙色高对比发光有机贴片 */
            .badge-hours {{
                background-color: rgba(249, 115, 22, 0.09);
                color: #ea580c;
                border: 1px solid rgba(249, 115, 22, 0.25);
                border-radius: 12px; 
                font-family: "PingFang SC", sans-serif;
                font-size: 10.5px;
                font-weight: 800;
                box-shadow: 0 1.5px 4px rgba(249, 115, 22, 0.06);
            }}
            
            .badge-icon {{
                width: 12px;
                height: 12px;
                margin-right: 4px;
                display: inline-block;
                color: inherit;
            }}
            
            .text-accent {{ color: var(--accent); }}
            .font-bold {{ font-weight: 700; }}
            .font-mono {{ font-family: "PingFang SC", -apple-system, sans-serif; }}
            
            .footer-end {{
                display: flex;
                align-items: center;
                justify-content: center;
                gap: 12px;
                font-size: 10.5px;
                color: var(--text-muted);
                letter-spacing: 0.25em;
                margin-top: 24px;
                font-weight: 800;
                position: relative;
                z-index: 2;
                opacity: 0.7;
            }}
            
            .footer-end::before, .footer-end::after {{
                content: "";
                width: 4px;
                height: 4px;
                background: linear-gradient(135deg, #3b82f6, #8b5cf6);
                border-radius: 50%;
                box-shadow: 0 1px 3px rgba(59, 130, 246, 0.3);
            }}
            
            .unapplied-tag {{
                display: inline-block;
                background: rgba(139, 92, 246, 0.06); 
                color: #7c3aed;
                padding: 3px 7px;
                border-radius: 6px;
                font-size: 10.5px;
                font-weight: 700;
                border: 1px solid rgba(139, 92, 246, 0.15);
                white-space: nowrap;
                line-height: 1.25;
                box-shadow: 0 1px 2px rgba(139, 92, 246, 0.02);
            }}
            .unapplied-empty {{
                font-size: 14px;
                font-weight: 700;
                color: var(--text-muted);
                line-height: 1.4;
            }}
            
            .empty-cards-notice {{
                display: flex;
                align-items: center;
                justify-content: center;
                gap: 8px;
                background: rgba(255, 255, 255, 0.45);
                border: 1px dashed rgba(100, 116, 139, 0.15);
                border-radius: var(--radius-md);
                padding: 24px;
                color: var(--text-muted);
                font-size: 13px;
                font-weight: 600;
                letter-spacing: 0.05em;
            }}
            
            .empty-notice-icon {{
                width: 14px;
                height: 14px;
                opacity: 0.6;
            }}

            .footer {{
                text-align: center;
                margin-top: 30px;
                font-size: 11px;
                color: var(--text-muted);
                border-top: 1px solid var(--border-light);
                padding-top: 15px;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <div class="header-logo-wrapper">
                {LOGO_SVG}
            </div>
            <h1>江西办 {week_info} 技术安排</h1>
            <div class="header-sub">TECHNICAL ARRANGEMENTS</div>
        </div>
        

        

        
        {sections_html}
        
        <!-- 底部 END 标识 -->
        <div class="footer-end">- END -</div>
        
        <!-- 智能裁剪定位标记线 -->
        <div style="height: 2px; background-color: #0f172a; margin-top: 15px;"></div>
    </body>
    </html>
    """
    return html

def calculate_stats(df):
    """
    计算数据集统计指标，包括计划总数、预计总工时、客户数以及未申请技术的销售人数
    """
    stats = {}
    stats['total_count'] = len(df)
    
    # 算工时，使用 pd.to_numeric 代替手写转换
    total_hours = 0
    try:
        total_hours = pd.to_numeric(df['预计工时（h）'], errors='coerce').fillna(0).sum()
    except Exception:
        pass
    stats['total_hours'] = f"{total_hours:g}"
    
    # 算客户数
    stats['total_customers'] = df['客户名称'].nunique()
    # 算销售数
    stats['total_sales'] = df['销售'].nunique()
    
    stats['generation_time'] = datetime.datetime.now().strftime("%Y-%m-%d")
    
    # 江西办固定销售名单
    SALES_LIST = ["邓正春", "饶达琴", "段振华", "钱丽云", "詹文成", "伍斌", "吴刚", "张进", "贺欢"]
    
    # 计算未申请技术的销售
    applied_sales = set(str(x).strip() for x in df['销售'].dropna().unique() if str(x).strip())
    unapplied_sales = [name for name in SALES_LIST if name not in applied_sales]
    
    stats['unapplied_sales_count'] = len(unapplied_sales)
    stats['unapplied_sales_names'] = "  ·  ".join(unapplied_sales)
    if unapplied_sales:
        stats['unapplied_sales_html'] = "".join([f'<span class="unapplied-tag">{name}</span>' for name in unapplied_sales])
    else:
        stats['unapplied_sales_html'] = '<span class="unapplied-empty">无</span>'
    
    return stats

def main():
    args = parse_args()
    
    input_path = os.path.abspath(args.input)
    if not os.path.exists(input_path):
        print(f"错误：找不到文件 {input_path}")
        sys.exit(1)
        
    output_dir = args.output_dir
    if not output_dir:
        output_dir = os.path.dirname(input_path)
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    
    # 1. 数据清理与加工
    print("正在加载和处理数据...")
    df = clean_and_prepare_data(input_path)
    
    # 提取所在月周作为顶部标题
    week_info = ""
    if '所在月周' in df.columns:
        non_empty_weeks = df[df['所在月周'].astype(str).str.strip() != '']['所在月周']
        if not non_empty_weeks.empty:
            week_info = str(non_empty_weeks.iloc[0]).strip()
    if not week_info:
        week_info = "本周"
    
    # 2. 导出整理后的 Excel
    excel_out = os.path.join(output_dir, f"{base_name}_整理.xlsx")
    print(f"正在导出整理后的 Excel：{excel_out}")
    df.to_excel(excel_out, index=False)
    beautify_excel(excel_out)
    
    # 3. 计算指标
    stats = calculate_stats(df)
    
    # 4. 生成临时 HTML 文件并截图
    temp_dir = os.path.join(output_dir, ".tmp_render")
    os.makedirs(temp_dir, exist_ok=True)
    
    card_html_path = os.path.join(temp_dir, "card_view.html")
    card_png_path = os.path.join(output_dir, f"{base_name}_手机卡片流.png")
    card_content = generate_card_html(df, stats, week_info)
    with open(card_html_path, "w", encoding="utf-8") as f:
        f.write(card_content)
        
    # 5. 调用 Headless Chrome 进行截图
    chrome_path = find_chrome_binary()
    if not chrome_path:
        print("错误：未能在系统中找到 Chrome/Chromium 浏览器。请安装 Chrome 或设置 CHROME_PATH 环境变量。")
        sys.exit(1)
    
    # 动态估算渲染高度，自适应调整视口，防截断并提升小图性能
    if df.empty:
        card_height = 800
    else:
        # 高精度动态高度计算：
        # 1. 顶部基础区域（header约120px + spacing约100px）
        total_height = 220
        
        # 2. 构造与渲染相同的所有销售列表，确保估算绝对准确
        SALES_LIST = ["邓正春", "饶达琴", "段振华", "钱丽云", "詹文成", "伍斌", "吴刚", "张进", "贺欢"]
        all_sales = list(SALES_LIST)
        if '销售' in df.columns:
            excel_sales = [str(x).strip() for x in df['销售'].dropna().unique() if str(x).strip()]
            for s in excel_sales:
                if s not in all_sales:
                    all_sales.append(s)

        # 3. 遍历每个销售估算其对应区块高度
        for s_name in all_sales:
            # 每一个销售 header 占约 45px
            total_height += 45
            
            if s_name in df['销售'].values:
                group = df[df['销售'] == s_name]
                for _, row in group.iterrows():
                    # 卡片基础高度（不含备注）：
                    # 包含卡片边距、Header（客户名称、兵种标签、预计工时）、Body网格（工作类型、协作人员、计划时间）
                    card_base = 110
                    
                    # 如果客户名称过长可能折行（超过 12 个字，累加 20px）
                    cust_name = str(row.get('客户名称', '')).strip()
                    if len(cust_name) > 12:
                        card_base += 20
                        
                    # 备注高度计算
                    remark = str(row.get('备注', '')).strip()
                    if remark:
                        lines = (len(remark) // 28) + 1
                        card_base += 25 + (lines * 20)
                        
                    # 卡片底部 margin 约 8px
                    total_height += card_base + 8
            else:
                # 任务数为 0 的销售渲染空安排占位卡，高度约 65px + 8px margin
                total_height += 73
            
        # 4. 底部 END 标识和间距约 100px，外加 300px 的绝对安全裕量（多余的部分会被 marker 扫描裁剪掉，确保绝对不截断）
        estimated_height = total_height + 100 + 300
        card_height = max(1100, min(20000, estimated_height))
    
    print("正在调用 Chrome Headless 渲染图片...")
    print(f"正在渲染手机卡片流长图 (480x{card_height} 自动裁剪) -> {card_png_path}")
    from pathlib import Path
    import tempfile
    import shutil
    
    card_html_url = Path(card_html_path).as_uri()
    
    # 每次运行生成唯一的随机临时 Profile 目录，彻底避免多次运行时的占锁冲突卡死
    chrome_profile_dir = tempfile.mkdtemp(prefix="chrome_profile_")
    
    cmd_card = [
        chrome_path,
        "--headless=new",                     # 启用 Chrome v109+ 新版极速 Headless 引擎，渲染与截图效率成倍提升
        "--disable-gpu",
        "--no-first-run",                   # 跳过首次运行欢迎引导，防止新建 Profile 时卡死
        "--no-default-browser-check",        # 跳过默认浏览器检测
        "--disable-background-networking",   # 禁用后台网络交互，加速启动并防无网环境卡死
        "--disable-sync",                    # 禁用账户同步
        "--disable-translate",               # 禁用翻译组件
        "--disable-default-apps",            # 禁用默认应用
        "--mute-audio",                      # 禁用音频
        "--disable-extensions",               # 禁用浏览器插件，缩短冷启动与进程加载时间
        "--disable-dev-shm-usage",            # 禁用共享内存文件系统限制，消除 IO 阻塞提速
        "--force-device-scale-factor=3",     # 提升为 3 倍视网膜缩放因子，提供极高清晰度无损画质
        f"--screenshot={card_png_path}",
        f"--window-size=480,{card_height}",
        f"--user-data-dir={chrome_profile_dir}",  # 隔离独立的临时配置目录，保证多实例和前后次运行绝对不锁死
        card_html_url
    ]
    try:
        print("正在启动 Chrome 进行截图...")
        result = subprocess.run(cmd_card, capture_output=True, text=True, timeout=90)
        if result.returncode != 0:
            print(f"Chrome 运行失败，退出码: {result.returncode}")
            print(f"Stdout:\n{result.stdout}")
            print(f"Stderr:\n{result.stderr}")
        else:
            print("Chrome 截图成功！")
    except subprocess.TimeoutExpired as e:
        print("Chrome 截图超时（90秒限制）！已强制终止。")
        print(f"目前已有的 stdout:\n{e.stdout}")
        print(f"目前已有的 stderr:\n{e.stderr}")
    except Exception as e:
        print(f"调用 Chrome 截图时发生未知错误: {e}")
    
    # 清理临时 Profile 目录
    try:
        shutil.rmtree(chrome_profile_dir)
    except Exception:
        pass
    
    # 执行智能高精度高度裁剪
    crop_image_by_marker(card_png_path)
    
    # 额外导出超高清手机分享版 JPG 格式，防止被微信二次压缩
    card_jpg_path = card_png_path.replace(".png", ".jpg")
    try:
        img = Image.open(card_png_path)
        img.convert("RGB").save(card_jpg_path, "JPEG", quality=95, subsampling=0)
        has_jpg = True
    except Exception as e:
        print(f"导出分享版 JPG 失败: {e}")
        has_jpg = False
    
    # 清理临时渲染 HTML
    try:
        os.remove(card_html_path)
        os.rmdir(temp_dir)
    except Exception:
        pass
        
    # 仅保留手机卡片流.jpg，在此清理作为中间产物的 PNG 图片
    try:
        if os.path.exists(card_png_path):
            os.remove(card_png_path)
    except Exception as e:
        print(f"清理临时 PNG 失败: {e}")
        
    print("\n转换全部完成！已成功生成以下文件：")
    print(f"1. [整理后的 Excel] {excel_out}")
    if has_jpg:
        print(f"2. [手机卡片流长图(JPG分享)] {card_jpg_path} (已优化体积，防社交软件二次压缩发虚)")

if __name__ == "__main__":
    main()
